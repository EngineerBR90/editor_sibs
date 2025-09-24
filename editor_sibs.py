import streamlit as st
import pandas as pd
import io
import zipfile
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

st.set_page_config(page_title="Limpador SIBS", layout="wide")

# ---------- Funções de processamento (adaptadas do seu protótipo) ----------

def limpar_arquivo_bytesio(uploaded_file):
    """Recebe um UploadedFile (BytesIO) e retorna um DataFrame limpo ou None"""
    try:
        name = uploaded_file.name
        # Escolher engine conforme extensão
        if name.lower().endswith('.xls'):
            df = pd.read_excel(uploaded_file, sheet_name=0, header=None, engine='xlrd')
        else:
            df = pd.read_excel(uploaded_file, sheet_name=0, header=None, engine='openpyxl')

        # Encontrar cabeçalho rapidamente
        cabecalho_linha = None
        for i in range(min(30, len(df))):
            if (pd.notna(df.iloc[i, 0]) and str(df.iloc[i, 0]).strip() == 'Código'):
                cabecalho_linha = i
                break

        if cabecalho_linha is None:
            return None

        dados = []
        for i in range(cabecalho_linha + 1, len(df)):
            linha = df.iloc[i]
            if pd.notna(linha[0]) and 'Total' in str(linha[0]):
                break
            if (pd.notna(linha[4]) and pd.notna(linha[12]) and pd.notna(linha[19])):
                try:
                    dados.append({
                        'Quantidade': float(linha[12]) / 10000,
                        'Item': str(linha[4]).strip(),
                        'Valor unitário [R$]': float(linha[17]) if pd.notna(linha[17]) else 0,
                        'Valor total [R$]': float(linha[19])
                    })
                except (ValueError, TypeError):
                    continue

        if not dados:
            return None

        return pd.DataFrame(dados)

    except Exception as e:
        st.error(f"Erro ao ler {uploaded_file.name}: {e}")
        return None


def aplicar_formatacao_bytes(buffer_bytes, aplicar_filtros=False):
    """Recebe bytes de um arquivo .xlsx em memória, aplica formatação simples e retorna bytes atualizados.
    Se aplicar_filtros=True, também cria um AutoFilter na planilha 'Faturamento' (caso exista)."""
    try:
        in_mem = io.BytesIO(buffer_bytes)
        workbook = load_workbook(filename=in_mem)

        # Aplica formatação básica em todas as planilhas dataframes (geralmente 'Todos' e/ou 'Faturamento')
        for ws in workbook.worksheets:
            # Formatação mínima: formatos numéricos
            for row in range(2, ws.max_row + 1):
                try:
                    ws[f'A{row}'].number_format = '0.00'
                except Exception:
                    pass
                try:
                    ws[f'C{row}'].number_format = '"R$" #,##0.00'
                except Exception:
                    pass
                try:
                    ws[f'D{row}'].number_format = '"R$" #,##0.00'
                except Exception:
                    pass

            # Ajuste de largura simples
            for col in ['A', 'B', 'C', 'D']:
                try:
                    ws.column_dimensions[col].width = 20
                except Exception:
                    pass

            # Se for a planilha de faturamento e o usuário requisitou filtros, cria AutoFilter na linha 1
            if aplicar_filtros and ws.title.lower().startswith('fatur'):
                try:
                    ws.auto_filter.ref = f"A1:D{ws.max_row}"
                except Exception:
                    pass

            # Se não for planilha de faturamento, adiciona uma soma rápida ao final (apenas na primeira planilha)
        try:
            ws0 = workbook.worksheets[0]
            ultima_linha = ws0.max_row
            soma_linha = ultima_linha + 1
            ws0[f'C{soma_linha}'] = 'Total'
            ws0[f'C{soma_linha}'].font = Font(bold=True)
            ws0[f'D{soma_linha}'] = f'=SUM(D2:D{ultima_linha})'
            ws0[f'D{soma_linha}'].font = Font(bold=True)
            ws0[f'D{soma_linha}'].number_format = '"R$" #,##0.00'
        except Exception:
            pass

        out_mem = io.BytesIO()
        workbook.save(out_mem)
        out_mem.seek(0)
        return out_mem.read()

    except Exception as e:
        st.warning(f"Formatação falhou: {e}")
        return buffer_bytes


def salvar_dataframe_para_bytes(df, aplicar_formatacao=True, aplicar_filtros=False):
    """Salva um DataFrame para bytes (.xlsx).
    Se aplicar_filtros=True, cria duas planilhas: 'Todos' (com todos os registros) e 'Faturamento' (apenas linhas onde Item contém 'mil' ou 'sod').
    Essa abordagem mantém todos os dados enquanto oferece uma aba já preparada para faturamento direto.
    """
    buf = io.BytesIO()

    if aplicar_filtros:
        # Escrever duas sheets: Todos e Faturamento (filtrada)
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Todos', index=False)
            # Filtrar por 'mil' ou 'sod' em Item (coluna B)
            try:
                mask = df['Item'].astype(str).str.contains(r'mil|sod', case=False, na=False)
                df_filtered = df[mask]
            except Exception:
                df_filtered = df.iloc[0:0]  # vazio, se algo falhar
            df_filtered.to_excel(writer, sheet_name='Faturamento', index=False)
        buf.seek(0)
        data = buf.read()

        if aplicar_formatacao:
            data = aplicar_formatacao_bytes(data, aplicar_filtros=True)

        return data

    else:
        df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)
        data = buf.read()

        if aplicar_formatacao:
            data = aplicar_formatacao_bytes(data, aplicar_filtros=False)

        return data


# ---------- UI Streamlit ----------

st.title("Limpador de planilhas SIBS")
st.markdown("Arraste arquivos `.xls` ou `.xlsx` para cá, ou selecione pela caixa abaixo. O app retorna os arquivos já limpos prontos para download.")

with st.expander("Instruções rápidas"):
    st.write("- Você pode enviar vários arquivos ao mesmo tempo.
- Se enviar mais de um arquivo, o download será entregue em um ZIP.
- Os arquivos resultantes terão o sufixo `_LIMPO.xlsx`.
- Use a opção 'Aplicar filtros para faturamento direto?' se quiser uma aba específica com apenas os itens que contenham 'mil' ou 'sod' no campo Item.")

uploaded_files = st.file_uploader("Escolha os arquivos (.xls / .xlsx)", type=['xls', 'xlsx'], accept_multiple_files=True)

col1, col2 = st.columns([3,1])
with col2:
    aplicar_filtros = st.checkbox("Aplicar filtros para faturamento direto?", value=False)
    btn_processar = st.button("Processar arquivos")

if btn_processar and uploaded_files:
    resultados = []
    progresso = st.progress(0)
    status = st.empty()

    start_total = time.time()
    for idx, up in enumerate(uploaded_files, start=1):
        status.info(f"Processando {up.name} ({idx}/{len(uploaded_files)})...")
        t0 = time.time()
        df_limpo = limpar_arquivo_bytesio(up)
        if df_limpo is None or df_limpo.empty:
            status.error(f"Falha ao processar {up.name} - formato inesperado ou ausência de dados detectada.")
            continue

        # Gerar bytes do arquivo final. Sempre aplicamos formatação mínima automaticamente.
        data_bytes = salvar_dataframe_para_bytes(df_limpo, aplicar_formatacao=True, aplicar_filtros=aplicar_filtros)
        filename_out = up.name.rsplit('.', 1)[0] + '_LIMPO.xlsx'
        resultados.append((filename_out, data_bytes, df_limpo))

        progresso.progress(int((idx/len(uploaded_files))*100))
        elapsed = time.time() - t0
        status.success(f"Concluído {up.name} em {elapsed:.2f}s - {len(df_limpo)} itens")

    total_elapsed = time.time() - start_total
    st.success(f"Processamento finalizado em {total_elapsed:.2f}s - {len(resultados)} arquivo(s) prontos")

    if not resultados:
        st.error("Nenhum arquivo processado com sucesso.")
    else:
        if len(resultados) == 1:
            nome, data_bytes, df_preview = resultados[0]
            st.download_button(label=f"Download: {nome}", data=data_bytes, file_name=nome, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            st.subheader("Pré-visualização (primeiras linhas)")
            st.dataframe(df_preview.head(20))
        else:
            # Criar ZIP em memória
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
                for nome, data_bytes, df_preview in resultados:
                    zf.writestr(nome, data_bytes)
            zip_buf.seek(0)
            st.download_button("Baixar todos os arquivos (.zip)", data=zip_buf.getvalue(), file_name='SIBS_LIMPAS.zip', mime='application/zip')

else:
    if uploaded_files:
        st.info(f"{len(uploaded_files)} arquivo(s) carregado(s). Clique em 'Processar arquivos' para começar.")
    else:
        st.info("Nenhum arquivo carregado ainda.")

# Fim do arquivo
